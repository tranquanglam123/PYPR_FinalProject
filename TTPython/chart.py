from PyQt5.QtWidgets import QMainWindow 
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication,QWidget,QVBoxLayout
from PyQt5.QtChart import QChart, QChartView, QBarSet, QPercentBarSeries, QBarCategoryAxis,QLineSeries,QScatterSeries,QBarSeries,QAbstractBarSeries
import sys
from PyQt5.QtGui import QPainter
from PyQt5.QtCore import Qt 
from openpyxl import *
#from sheet_data import DEVICE_DATA
#import time


class Barchart():    
    def __init__(self):
        #----------------- Khởi tạo các thuộc tính ------------------
        self.file = None
        self.data = []
        #tạo thuộc tính set chứa dữ liệu của biểu đồ
        self.set = []
        
 
        #Tạo biểu đồ cột theo phàn trăm

        self.series = QPercentBarSeries()
        
 
        #Tạo biểu đồ, thêm series vào biểu đồ
        self.chart = QChart()
        self.chart.addSeries(self.series)
        self.chart.setTitle("Barchart Percent Example")
        self.chart.setAnimationOptions(QChart.SeriesAnimations)
        self.chart.setTheme(QChart.ChartThemeDark)
 
 
        #Tạo các cột(trục) cho biểu đò
        self.categories = []
        self.axis = QBarCategoryAxis()
        self.axis.append(self.categories)
        self.chart.createDefaultAxes()
        self.chart.setAxisX(self.axis, self.series)

        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignBottom)    
        
 
        #Tạo đồ thị, thêm biểu đồ vào đồ thị 
        self.chartview = QChartView(self.chart)
 
        self.vbox = QVBoxLayout() #thiết lập layout 
        self.vbox.addWidget(self.chartview)#thêm Widget đồ thị
    def update_data(self):
        self.series = QPercentBarSeries()
        self.get_data()
        for i in self.data:
            print(i)
            try:
                self.series.append(float(i[0]),float(i[1])) #bắt lỗidữ  liệu k đúng ctrnih vẫn chạy          
                #print(self.series.value)
            except:
                continue    
    def get_data(self):
        #self.filename = QFileDialog.getOpenFileName(filter="CSV (*.csv)")[0]
        #self.df = pd.read_csv(self.filename, encoding='utf-8').fillna(0)

        workbook = load_workbook(filename = self.file)
        sheet = workbook.worksheets[2]          #sheet 3: dữ liệu của bar chart
        # print(sheet)
        self.categories = []
        self.set = []
        self.series = QPercentBarSeries()
        #print(self.categories)
        
        for row in sheet.rows:
            print(row[0].value,"--")
            if row[0].value == None:
                for cell in row:
                    self.categories.append(cell.value)
            else:
                temp = QBarSet(row[0].value)
                for i in range(1,len(row)):
                    temp << row[i].value
                self.set.append(temp)

        #Vòng lặp không cần thiết, thay cho hàm update_data
        for obj in self.set:
            self.series.append(obj)
        #Fixed
        self.axis.append(self.categories)

    def assign_chart(self):
        self.widget.setLayout(self.vbox)
class Histogramchart(Barchart):
    def __init__(self):
        #----------------- Khởi tạo các thuộc tính ------------------
        self.file = None
        self.data = []
        #Tạo barseries
        self.set = []
        #Tạo biểu đồ cột theo phần trăm

        self.series = QBarSeries()
        #Tạo biểu đồ và thêm series
        self.chart = QChart()
        self.chart.addSeries(self.series)
        self.chart.setTitle("HISTOGRAM !!")
        self.chart.setAnimationOptions(QChart.SeriesAnimations)
        self.chart.setTheme(QChart.ChartThemeDark)
  
        #Tạo cột(trục) cho biểu đồ
        self.categories = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
 
        self.axis = QBarCategoryAxis()
        self.axis.append(self.categories)
        self.chart.createDefaultAxes()
        self.chart.setAxisX(self.axis, self.series)
  
        self.series.setBarWidth(6.1)
 
        #Tạo đồ thị, thêm biểu đồ vào
        self.chartview = QChartView(self.chart)
 
        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.chartview)
    def update_data(self):
        self.series = QBarSeries()
        self.get_data()

        for i in self.data:
            try:
                self.series.append(float(i[0]),float(i[1]))
                #print(self.series.value)
            except:
                continue    


    def get_data(self):
        workbook = load_workbook(filename = self.file)
        sheet = workbook.worksheets[2]          #sheet 3
        self.categories = []
        self.set = []
        self.series = QBarSeries()
        #print(self.categories)
        
        for row in sheet.rows:
            print(row[0].value,"--")
            if row[0].value == None:
                for cell in row:
                    self.categories.append(cell.value)
            else:
                temp = QBarSet(row[0].value)
                for i in range(1,len(row)):
                    temp << row[i].value
                self.set.append(temp)

        for obj in self.set:
            self.series.append(obj)
        self.axis.append(self.categories)
        pass



class Linechart():
    def __init__(self):
        #----------------- Khởi tạo các thuộc tính ------------------
        self.file = None
        self.data = []
        #tạo thuộc tính series chứa dữ liệu của biểu đồ
        self.series = QLineSeries()
        
        '''
        self.series << QPointF(11,1) << QPointF(13,3)\
        << QPointF(17,6) << QPointF(18,3) << QPointF(20,20)
        '''
 
        self.chart = QChart()
        self.chart.addSeries(self.series)
        self.chart.setAnimationOptions(QChart.SeriesAnimations)
        self.chart.setTitle("Line Chart")
        self.chart.setTheme(QChart.ChartThemeBlueCerulean)
        self.chart.createDefaultAxes()
 
        self.chartview = QChartView(self.chart)
        #Render Hint là hướng dãn render, hiển thị hình ảnh
        #Render Hint ở dưới là antialiasing, giúp cho khi hiển thị chart không bị góc cạnh(khử răng cưa)
        self.chartview.setRenderHint(QPainter.Antialiasing) 
        

        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.chartview)
    def update_data(self):
        self.series = QLineSeries()
        print(self.series)
        self.get_data()
        # print(self.data)
        for i in self.data:
            try:
                self.series.append(float(i[0]),float(i[1]))
                #print(self.series.value)
            except:
                continue    
    def get_data(self):
        workbook = load_workbook(filename = self.file)
        sheet = workbook.worksheets[0]          #sheet 1 
        #print(sheet)
        self.data=[]
        for row in sheet.rows:
            temp = []
            for cell in row:
                # print(cell.value)
                temp.append(cell.value)
                #print(self.data)
            self.data.append(temp)
                

    def assign_chart(self,widget):
        self.widget.setLayout(self.vbox)

class Scatterchart():
    def __init__(self):
        #self.data = DEVICE_DATA()


        #----------------- Khởi tạo các thuộc tính ------------------
        self.file = None
        self.data = []
        self.series = QScatterSeries()

        # self.series << QPointF(11,1) << QPointF(13,3)\
        # << QPointF(17,6) << QPointF(18,3) << QPointF(20,20)
 
 
        self.chart = QChart()
        self.chart.addSeries(self.series)
        self.chart.setAnimationOptions(QChart.SeriesAnimations)
        self.chart.setTitle("Relative Moisture to Temperature Ratio")
        self.chart.setTheme(QChart.ChartThemeDark)
        self.chart.createDefaultAxes()
 
        self.chartview = QChartView(self.chart)
        self.chartview.setRenderHint(QPainter.Antialiasing)
        

        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.chartview)
    def update_data(self):
        self.series = QScatterSeries()
        self.get_data()
        # print(self.data)
        for i in self.data:
            try:
                self.series.append(float(i[0]),float(i[1]))
                #print(self.series.value)
            except:
                continue    
    def get_data(self):
        workbook = load_workbook(filename = self.file)
        sheet = workbook.worksheets[0]          #sheet 1 
        # print(sheet)
        self.data=[]
        for row in sheet.rows:
            temp = []
            for cell in row:
                # print(cell.value)
                temp.append(cell.value)
                #print(self.data)
            self.data.append(temp)
                

    def assign_chart(self,widget):
        self.widget.setLayout(self.vbox)
if __name__ == "__main__":           
    App = QApplication(sys.argv)
    #window = Window()
    #window.show()
    menu = Barchart()
    menu.assign_chart()
    sys.exit(App.exec())